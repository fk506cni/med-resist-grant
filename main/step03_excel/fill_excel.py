#!/usr/bin/env python3
"""fill_excel.py — 様式6, 7, 8 Excel記入スクリプト

data/source/ のExcel様式ファイルをコピーし、YAML データに基づいて
セルに値を書き込む。書式（フォント、罫線、列幅、行高、セル結合、
ドロップダウンバリデーション）はすべて維持する。

Usage:
    python fill_excel.py \\
        --config main/00_setup/config.yaml \\
        --researchers main/00_setup/researchers.yaml \\
        --source-dir data/source \\
        --output main/step03_excel/output/
"""

import argparse
import io
import re
import shutil
import sys
import warnings
import zipfile
from collections import OrderedDict
from pathlib import Path

import yaml


# ============================================================================
# YAML helpers
# ============================================================================

def load_yaml(path):
    """Load YAML file."""
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ============================================================================
# extLst preservation
# ============================================================================
# openpyxl strips x14 data validation extensions (the "Data Validation
# extension is not supported" warning).  To preserve dropdown lists that use
# the newer x14 format we extract the <extLst> block from the original xlsx
# ZIP and re-inject it after openpyxl has saved.

def _extract_sheet_xml_parts(xlsx_path, sheet_xml="xl/worksheets/sheet1.xml"):
    """Extract extLst block and namespace declarations from a worksheet XML.

    Returns (extlst_xml, extra_ns_attrs) where extra_ns_attrs is a dict of
    namespace declarations (e.g. 'xmlns:xr' → URI) and mc:Ignorable that are
    present in the original but stripped by openpyxl.
    """
    with zipfile.ZipFile(xlsx_path, "r") as z:
        content = z.read(sheet_xml).decode("utf-8")

    # Extract extLst
    extlst = None
    start = content.find("<extLst>")
    if start >= 0:
        end = content.find("</extLst>", start)
        if end >= 0:
            extlst = content[start : end + len("</extLst>")]

    # Extract namespace declarations from <worksheet ...> tag
    ws_start = content.find("<worksheet")
    ws_end = content.find(">", ws_start)
    ws_tag = content[ws_start : ws_end + 1]
    # Collect xmlns:*, mc:Ignorable, xr:uid attributes
    extra_attrs = {}
    for m in re.finditer(r'(xmlns:\w+)="([^"]*)"', ws_tag):
        extra_attrs[m.group(1)] = m.group(2)
    mc_match = re.search(r'mc:Ignorable="([^"]*)"', ws_tag)
    if mc_match:
        extra_attrs["mc:Ignorable"] = mc_match.group(1)
    xr_uid = re.search(r'xr:uid="([^"]*)"', ws_tag)
    if xr_uid:
        extra_attrs["xr:uid"] = xr_uid.group(1)

    return extlst, extra_attrs


def _reinject_extlst(xlsx_path, extlst_xml, extra_ns_attrs,
                     sheet_xml="xl/worksheets/sheet1.xml"):
    """Re-inject extLst and namespace declarations into the worksheet XML."""
    if extlst_xml is None:
        return
    buf = io.BytesIO()
    with zipfile.ZipFile(xlsx_path, "r") as z_in:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z_out:
            for item in z_in.infolist():
                data = z_in.read(item.filename)
                if item.filename == sheet_xml:
                    content = data.decode("utf-8")

                    # Add missing namespace declarations to <worksheet> tag
                    ws_match = re.search(r"<worksheet\b([^>]*)>", content)
                    if ws_match and extra_ns_attrs:
                        existing = ws_match.group(1)
                        additions = []
                        for attr, val in extra_ns_attrs.items():
                            if f'{attr}=' not in existing:
                                additions.append(f'{attr}="{val}"')
                        if additions:
                            insert = ws_match.start(1) + len(existing)
                            content = (
                                content[:insert]
                                + " " + " ".join(additions)
                                + content[insert:]
                            )

                    # Remove any existing (possibly truncated) extLst
                    ex_start = content.find("<extLst>")
                    if ex_start >= 0:
                        ex_end = content.find("</extLst>", ex_start)
                        if ex_end >= 0:
                            content = (
                                content[:ex_start]
                                + content[ex_end + len("</extLst>") :]
                            )

                    # Insert before </worksheet>
                    close_tag = "</worksheet>"
                    insert_pos = content.rfind(close_tag)
                    if insert_pos >= 0:
                        content = (
                            content[:insert_pos] + extlst_xml + content[insert_pos:]
                        )
                    data = content.encode("utf-8")
                z_out.writestr(item, data)
    with open(xlsx_path, "wb") as f:
        f.write(buf.getvalue())


# ============================================================================
# Budget helpers
# ============================================================================

def _yearly_total(year_data, indirect_rate):
    """Calculate total cost (direct + indirect) for a year, in 千円."""
    direct = (
        year_data.get("equipment", 0)
        + year_data.get("consumables", 0)
        + year_data.get("travel", 0)
        + year_data.get("personnel", 0)
        + year_data.get("other", 0)
    )
    return int(direct * (1 + indirect_rate))


# ============================================================================
# 様式6: 申請概要
# ============================================================================

def fill_youshiki6(config, researchers, source_path, output_path):
    """Fill 様式6 (application overview) — row 21."""
    import openpyxl

    # Preserve extLst (x14 data validations) before openpyxl strips them
    extlst, extra_ns = _extract_sheet_xml_parts(source_path)

    shutil.copy2(source_path, output_path)
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb["様式6"]

    # Build theme lookup from リスト sheet: number → display string
    ws_list = wb["リスト"]
    theme_map = {}
    for row in range(4, 50):
        val = ws_list.cell(row=row, column=1).value
        if val is None:
            break
        m = re.match(r"\((\d+)\)", val)
        if m:
            theme_map[int(m.group(1))] = val

    ROW = 21
    proj = config["project"]
    pi = researchers["pi"]
    lead = config["lead_institution"]
    subs = config.get("sub_institutions", [])
    budget = config["budget"]

    # D: 研究テーマ — dropdown value from リスト sheet
    ws.cell(row=ROW, column=4).value = theme_map.get(
        proj["theme_number"], f"({proj['theme_number']})"
    )

    # E: キーワード
    ws.cell(row=ROW, column=5).value = "・".join(proj["keywords"])

    # F: 研究分野
    ws.cell(row=ROW, column=6).value = proj["field"]

    # G: タイプ
    ws.cell(row=ROW, column=7).value = proj["type"]

    # H: 重複応募有無
    ws.cell(row=ROW, column=8).value = "無" if not proj["duplicate_application"] else "有"

    # I: 研究課題名（日）
    ws.cell(row=ROW, column=9).value = proj["title_ja"]

    # J: 研究代表者役職
    ws.cell(row=ROW, column=10).value = pi["position"]

    # K: 研究代表者名
    ws.cell(row=ROW, column=11).value = pi["name_ja"]

    # L: 代表研究機関
    ws.cell(row=ROW, column=12).value = lead["name"]

    # M: 分担研究機関（改行区切り）
    sub_names = [s["name"] for s in subs]
    ws.cell(row=ROW, column=13).value = "\n".join(sub_names) if sub_names else ""

    # N: 研究期間 — dropdown expects "3年" format
    ws.cell(row=ROW, column=14).value = f"{proj['period_years']}年"

    # O–S: 1–5年目の研究費（千円、直接+間接）
    indirect_rate = budget["indirect_rate"]
    for yr in budget["yearly"]:
        col = 14 + yr["year"]  # O=15 for year 1, P=16, ...
        ws.cell(row=ROW, column=col).value = _yearly_total(yr, indirect_rate)

    # T: 総額 — =SUM(O21:S21) already present; do not overwrite

    # U: メールアドレス
    ws.cell(row=ROW, column=21).value = pi["contact"]["email"]

    # V: 郵送先
    ws.cell(row=ROW, column=22).value = pi["contact"]["postal"]

    # W: 代表研究機関の種別
    ws.cell(row=ROW, column=23).value = lead["type"]

    # X: 中小企業
    ws.cell(row=ROW, column=24).value = "○" if lead["is_sme"] else "×"

    # Y: スタートアップ
    ws.cell(row=ROW, column=25).value = "○" if lead["is_startup"] else "×"

    # Z: 大学発スタートアップの大学名
    ws.cell(row=ROW, column=26).value = lead.get("university_origin") or "該当なし"

    # AA: 分担研究機関の中小企業名
    sme = [s["name"] for s in subs if s.get("is_sme")]
    ws.cell(row=ROW, column=27).value = "\n".join(sme) if sme else "該当なし"

    # AB: 分担研究機関のスタートアップ名
    startups = [s["name"] for s in subs if s.get("is_startup")]
    ws.cell(row=ROW, column=28).value = "\n".join(startups) if startups else "該当なし"

    # AC: 分担研究機関が大学発スタートアップの大学名
    uni_orig = [s["university_origin"] for s in subs if s.get("university_origin")]
    ws.cell(row=ROW, column=29).value = "\n".join(uni_orig) if uni_orig else "該当なし"

    wb.save(output_path)

    # Re-inject x14 data validations and namespace declarations
    _reinject_extlst(output_path, extlst, extra_ns)
    print(f"  様式6 → {output_path}")


# ============================================================================
# 様式7: 研究者の一覧
# ============================================================================

def fill_youshiki7(config, researchers, source_path, output_path):
    """Fill 様式7 (researcher list) — rows 24+."""
    import openpyxl

    shutil.copy2(source_path, output_path)
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb[wb.sheetnames[0]]  # 「採択課題抜粋」

    # Build researcher list grouped by institution (PI first)
    pi = researchers["pi"]
    co_invs = researchers.get("co_investigators", [])

    groups = OrderedDict()
    pi_inst = pi.get("affiliation", config["lead_institution"]["name"])
    groups.setdefault(pi_inst, []).append(pi)
    for ci in co_invs:
        inst = ci.get("institution", ci.get("affiliation", ""))
        groups.setdefault(inst, []).append(ci)

    DATA_START = 24  # row 23 = header, 24+ = data

    # Unmerge pre-existing merged cells in the data area (rows >= 24)
    to_unmerge = []
    for m in list(ws.merged_cells.ranges):
        cell_range = str(m)
        row_nums = [int(n) for n in re.findall(r"\d+", cell_range)]
        if any(r >= DATA_START for r in row_nums):
            to_unmerge.append(cell_range)
    for rng in to_unmerge:
        ws.unmerge_cells(rng)

    title = config["project"]["title_ja"]
    row = DATA_START
    inst_ranges = []  # [(start_row, end_row)]

    for inst_name, members in groups.items():
        start_row = row
        for member in members:
            # B: 研究課題名 (write in first data row only; merge spans all)
            if row == DATA_START:
                ws.cell(row=row, column=2).value = title

            # C: 研究機関名 (write in first row of each institution group)
            if row == start_row:
                ws.cell(row=row, column=3).value = inst_name

            # D: 研究者氏名
            ws.cell(row=row, column=4).value = member["name_ja"]

            # E: 部局・職/職階
            dept = member.get("department", "")
            pos = member.get("position", "")
            ws.cell(row=row, column=5).value = f"{dept}/{pos}" if dept else pos

            row += 1
        inst_ranges.append((start_row, row - 1))

    end_row = row - 1

    # Merge B (research title) across all data rows
    if end_row > DATA_START:
        ws.merge_cells(
            start_row=DATA_START, start_column=2, end_row=end_row, end_column=2
        )

    # Merge C (institution name) for each institution group
    for start, end in inst_ranges:
        if end > start:
            ws.merge_cells(start_row=start, start_column=3, end_row=end, end_column=3)

    wb.save(output_path)
    print(f"  様式7 → {output_path}")


# ============================================================================
# 様式8: 連絡先
# ============================================================================

def fill_youshiki8(config, researchers, source_path, output_path):
    """Fill 様式8 (contact emails) — rows 3–12."""
    import openpyxl

    shutil.copy2(source_path, output_path)
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb[wb.sheetnames[0]]  # Sheet1

    pi = researchers["pi"]
    emails = config.get("contacts", {}).get("emails", [])

    # B3: PI email (required)
    ws.cell(row=3, column=2).value = pi["contact"]["email"]

    # B4–B12: additional contact emails (skip first = PI)
    for i, email in enumerate(emails[1:], start=4):
        if i > 12:
            break
        ws.cell(row=i, column=2).value = email

    wb.save(output_path)
    print(f"  様式8 → {output_path}")


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Fill Excel forms (様式6, 7, 8)")
    parser.add_argument("--config", default="main/00_setup/config.yaml")
    parser.add_argument("--researchers", default="main/00_setup/researchers.yaml")
    parser.add_argument("--source-dir", default="data/source")
    parser.add_argument("--output", default="main/step03_excel/output/")
    args = parser.parse_args()

    config = load_yaml(args.config)
    researchers = load_yaml(args.researchers)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)
    source_dir = Path(args.source_dir)

    print("Filling Excel forms...")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)

        fill_youshiki6(
            config, researchers,
            source_dir / "r08youshiki6.xlsx",
            output_dir / "youshiki6.xlsx",
        )
        fill_youshiki7(
            config, researchers,
            source_dir / "r08youshiki7.xlsx",
            output_dir / "youshiki7.xlsx",
        )
        fill_youshiki8(
            config, researchers,
            source_dir / "r08youshiki8.xlsx",
            output_dir / "youshiki8.xlsx",
        )

    print("Done.")


if __name__ == "__main__":
    main()
